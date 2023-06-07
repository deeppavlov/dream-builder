import argparse
import csv
import logging
from pathlib import Path
from typing import Union

import openpyxl
from deeppavlov_dreamtools import list_dists, utils
from deeppavlov_dreamtools.distconfigs.components import DreamComponent

from database import enums

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s %(levelname)s: %(message)s")
handler = logging.StreamHandler()
handler.setFormatter(formatter)
logger.addHandler(handler)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("-d", "--dream", type=str, help="path to dream root directory")
    parser.add_argument("-i", "--initial-file", type=str, help="path to initial data .xlsx file")
    parser.add_argument(
        "-o", "--output-dir", type=str, help="directory where the initial database files will be created"
    )

    return parser.parse_args()


INITIAL_DATA_LISTS = [
    "google_user",
    "deployments",
    "virtual_assistant_names",
    "public_virtual_assistant_names",
    "lm_service",
    "api_key",
    "role",
]


def find_row(rows: list[dict], key: str, value: Union[str, int]):
    for row in rows:
        if row[key] == value:
            return row


def read_xlsx_file(filename: Union[Path, str]) -> dict:
    workbook = openpyxl.load_workbook(filename)

    sheet_names = workbook.sheetnames
    missing_data_lists = set(INITIAL_DATA_LISTS) - set(sheet_names)
    if missing_data_lists:
        raise ValueError(f"File {filename} is missing data list(s): {', '.join(missing_data_lists)}")

    data_dict = {}
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        header_row = [cell.value for cell in worksheet[1]]

        data_rows = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            data_rows.append(row)

        data_dict[sheet_name] = []
        for row in data_rows:
            row_dict = {}
            for col_num, cell_value in enumerate(row):
                row_dict[header_row[col_num]] = cell_value
            data_dict[sheet_name].append(row_dict)

    return data_dict


def run(dream_root: Union[Path, str], initial_file: Union[Path, str], output_dir: Union[Path, str]):
    dream_root = Path(dream_root)

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    all_services = {}
    all_components = {}

    initial_secret = read_xlsx_file(initial_file)

    for table_name in ["role", "google_user", "lm_service", "api_key"]:
        with open(f"{output_dir}/{table_name}.tsv", "w", encoding="utf-8") as table_tsv_f:
            table_csv_writer = csv.writer(table_tsv_f, delimiter="\t", quotechar='"', quoting=csv.QUOTE_MINIMAL)
            table_csv_writer.writerow(list(initial_secret[table_name][0].keys()))

            for row in initial_secret[table_name]:
                if any(list(row.values())):
                    table_csv_writer.writerow(list(row.values()))

    with (
        open(f"{output_dir}/component.tsv", "w", encoding="utf-8") as component_tsv_f,
        open(f"{output_dir}/service.tsv", "w", encoding="utf-8") as service_tsv_f,
    ):
        components_csv_writer = csv.writer(component_tsv_f, delimiter="\t", quotechar='"', quoting=csv.QUOTE_MINIMAL)
        components_csv_writer.writerow(
            [
                "source",
                "name",
                "display_name",
                # "container_name",
                "component_type",
                "model_type",
                "is_customizable",
                "author_id",
                "description",
                "ram_usage",
                "gpu_usage",
                # "port",
                "group",
                "service_id",
                "endpoint",
                # "build_args",
                "prompt",
                "prompt_goals",
                "lm_service_id",
                "date_created",
            ]
        )
        services_csv_writer = csv.writer(service_tsv_f, delimiter="\t", quotechar='"', quoting=csv.QUOTE_MINIMAL)
        services_csv_writer.writerow(
            [
                "source",
                "name",
            ]
        )

        current_service_id = 0
        current_component_id = 1

        for component_path in (dream_root / "components").rglob("*.yml"):
            component = utils.load_yml(component_path)
            service_card = utils.load_yml(dream_root / component["service"] / "service.yml")
            service_env_card = utils.load_yml(dream_root / component["service"] / "environment.yml")

            if component["service"] not in all_services:
                current_service_id += 1
                all_services[component["service"]] = current_service_id
                services_csv_writer.writerow(
                    [
                        component["service"],
                        service_card["name"],
                    ]
                )

            prompt = prompt_goals = None
            lm_service_id = None
            if component["name"].endswith("prompted_skill"):

                if service_env_card and service_env_card.get("GENERATIVE_SERVICE_URL"):
                    lm_service, _, _ = utils.parse_connector_url(service_env_card["GENERATIVE_SERVICE_URL"])
                    lm_service_row = find_row(initial_secret["lm_service"], key="name", value=lm_service)
                    if lm_service_row:
                        if not lm_service_row["is_hosted"]:
                            logger.error(
                                f"{component_path} is using LLM service {lm_service_row['name']} "
                                "which is no longer hosted!"
                            )
                        elif not lm_service_row["is_maintained"]:
                            logger.warning(
                                f"{component_path} is using LLM service {lm_service_row['name']} "
                                "which is no longer maintained!"
                            )
                        lm_service_id = lm_service_row["id"]

                    prompt_file = service_env_card["PROMPT_FILE"]
                    prompt_data = utils.load_json(dream_root / prompt_file)
                    prompt, prompt_goals = prompt_data["prompt"], prompt_data["goals"]

            all_components[str(component_path.relative_to(dream_root))] = current_component_id
            current_component_id += 1

            components_csv_writer.writerow(
                [
                    str(component_path.relative_to(dream_root)),
                    component["name"],
                    component["display_name"],
                    # component["container_name"],
                    component["component_type"],
                    component["model_type"],
                    int(component["is_customizable"]),
                    1,
                    component["description"],
                    component["ram_usage"],
                    component["gpu_usage"],
                    # component["port"],
                    component["group"],
                    all_services[component["service"]],
                    component["endpoint"],
                    prompt,
                    prompt_goals,
                    lm_service_id,
                    component["date_created"],
                ]
            )

    with (
        open(f"{output_dir}/virtual_assistant.tsv", "w", encoding="utf-8") as va_tsv_f,
        open(f"{output_dir}/virtual_assistant_component.tsv", "w", encoding="utf-8") as va_components_tsv_f,
        open(f"{output_dir}/publish_request.tsv", "w", encoding="utf-8") as publish_tsv_f,
        open(f"{output_dir}/deployment.tsv", "w", encoding="utf-8") as deployment_tsv_f,
    ):
        va_csv_writer = csv.writer(va_tsv_f, delimiter="\t", quotechar='"')
        va_csv_writer.writerow(
            ["author_id", "source", "name", "display_name", "description", "private_visibility", "date_created"]
        )

        va_components_csv_writer = csv.writer(va_components_tsv_f, delimiter="\t", quotechar='"')
        va_components_csv_writer.writerow(["virtual_assistant_id", "component_id", "is_enabled"])

        publish_csv_writer = csv.writer(publish_tsv_f, delimiter="\t", quotechar='"')
        publish_csv_writer.writerow(
            ["slug", "virtual_assistant_id", "user_id", "public_visibility", "state", "reviewed_by_user_id"]
        )

        deployment_csv_writer = csv.writer(deployment_tsv_f, delimiter="\t", quotechar='"')
        deployment_csv_writer.writerow(["virtual_assistant_id", "chat_host", "chat_port", "state"])

        current_assistant_dist_id = 1
        for dist in list_dists(dream_root):
            if dist.name in [row["name"] for row in initial_secret["virtual_assistant_names"]]:
                va_row = [
                    1,
                    f"assistant_dists/{dist.dist_path.name}",
                    dist.name,
                    dist.pipeline.metadata.display_name,
                    dist.pipeline.metadata.description,
                    enums.VirtualAssistantPrivateVisibility.UNLISTED_LINK.value,
                    dist.pipeline.metadata.date_created.strftime("%Y-%m-%dT%H:%M:%S"),
                ]
                va_csv_writer.writerow(va_row)

                dist_components = {
                    "last_chance_service": dist.pipeline.last_chance_service,
                    "timeout_service": dist.pipeline.timeout_service,
                    "annotators": dist.pipeline.annotators,
                    "response_annotators": dist.pipeline.response_annotators,
                    "response_annotator_selectors": dist.pipeline.response_annotator_selectors,
                    "candidate_annotators": dist.pipeline.candidate_annotators,
                    "skill_selectors": dist.pipeline.skill_selectors,
                    "skills": dist.pipeline.skills,
                    "response_selectors": dist.pipeline.response_selectors,
                }

                for group_name, component_group in dist_components.items():
                    if isinstance(component_group, DreamComponent):
                        va_components_row = [
                            current_assistant_dist_id,
                            all_components[str(component_group.component_file)],
                            1,
                        ]
                        va_components_csv_writer.writerow(va_components_row)
                    elif isinstance(component_group, dict):
                        for component_name, component_value in component_group.items():
                            va_components_row = [
                                current_assistant_dist_id,
                                all_components[str(component_value.component_file)],
                                1,
                            ]
                            va_components_csv_writer.writerow(va_components_row)

                publish_row = [
                    dist.name,
                    current_assistant_dist_id,
                    1,
                    enums.VirtualAssistantPublicVisibility.PUBLIC_TEMPLATE.value,
                    enums.PublishRequestState.APPROVED.value,
                    1,
                ]
                publish_csv_writer.writerow(publish_row)

                deployment = find_row(initial_secret["deployments"], key="virtual_assistant_name", value=dist.name)
                if deployment:
                    deployment_row = [
                        current_assistant_dist_id,
                        deployment["host"],
                        deployment["port"],
                        "UP",
                    ]
                    deployment_csv_writer.writerow(deployment_row)

                current_assistant_dist_id += 1


if __name__ == "__main__":
    args = parse_args()

    logger.info(f"Generating initial data\nDream path: {args.dream}\nSecret settings file: {args.initial_file}")
    run(args.dream, args.initial_file, args.output_dir)
    logger.info(f"Finished creating initial data tsv files in {args.output_dir}")
